"""Excel export of an event's results.

Produces a multi-sheet workbook mirroring the manual results spreadsheet:

- an **Overall** sheet: one row per team, one column per checkpoint (in visit
  order) holding that checkpoint's score, plus the team's total; and
- one **Checkpoint N** sheet per checkpoint, breaking each team's result down
  into match points, extra shots, penalties and notes.

The export is a plain read of persisted ``ActivityResult`` rows scoped to the
event; it never recomputes scores.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.activity import ActivityResult
from app.models.checkpoint import CheckPoint
from app.models.team import Team
from app.services.event_results_query import (
    EventResultsData,
    EventResultsQuery,
    result_notes,
    result_penalty,
)

# Built-in penalty keys in ActivityResult.penalties (staff evaluation form).
# Stored as accumulated positive magnitudes.
_VOMIT_KEY = "vomit"
_NOT_DRINKING_KEY = "not_drinking"

_HEADER_FILL = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFFFF")
_BODY_FONT = Font(name="Arial")
_CENTER = Alignment(horizontal="center")


def _style_header(ws: Worksheet, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER


def _autosize(ws: Worksheet, ncols: int, min_width: int = 12) -> None:
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        longest = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)),
            default=min_width,
        )
        ws.column_dimensions[letter].width = max(min_width, longest + 2)


class ExportService:
    """Builds the results workbook for a single event."""

    def __init__(self, db: AsyncSession):
        self.db = db
        self._query = EventResultsQuery(db)

    async def _teams(self, event_id: int) -> list[Team]:
        return await self._query.teams(event_id)

    async def _checkpoints(self, event_id: int) -> list[CheckPoint]:
        return await self._query.checkpoints(event_id)

    async def _results(self, event_id: int) -> list[ActivityResult]:
        return await self._query.results(event_id)

    # Kept as instance-level aliases (rather than importing the shared
    # functions directly at call sites) so the existing unit tests, which
    # stub `_teams`/`_checkpoints`/`_results` on the instance, keep working
    # unchanged — see app/tests/unit/services/test_export_service.py.
    _penalty = staticmethod(result_penalty)
    _notes = staticmethod(result_notes)

    async def build_workbook(self, event_id: int) -> bytes:
        teams = await self._teams(event_id)
        checkpoints = await self._checkpoints(event_id)
        results = await self._results(event_id)
        data = EventResultsData(teams=teams, checkpoints=checkpoints, results=results)

        wb = Workbook()
        self._build_overall_sheet(wb, teams, checkpoints, data.opponent_of, data.cp_score)
        for idx, cp in enumerate(checkpoints, start=1):
            self._build_checkpoint_sheet(
                wb, cp, idx, teams, data.opponent_of, data.cp_result, data.cp_score
            )

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _build_overall_sheet(
        self,
        wb: Workbook,
        teams: list[Team],
        checkpoints: list[CheckPoint],
        opponent_of: dict[int, str],
        cp_score: dict[tuple[int, int], float],
    ) -> None:
        ws = wb.active
        ws.title = "Overall"

        headers = (
            ["Team", "Versus Pair"]
            + [f"Checkpoint {i}" for i in range(1, len(checkpoints) + 1)]
            + ["Total Points"]
        )
        ws.append(headers)

        for team in teams:
            row: list[Any] = [team.name, opponent_of.get(team.id, "")]
            total = 0.0
            for cp in checkpoints:
                score = cp_score.get((team.id, cp.id), 0.0)
                total += score
                row.append(score)
            row.append(total)
            ws.append(row)

        self._finalize(ws, len(headers), body_center_from_col=3)

    def _build_checkpoint_sheet(
        self,
        wb: Workbook,
        checkpoint: CheckPoint,
        index: int,
        teams: list[Team],
        opponent_of: dict[int, str],
        cp_result: dict[tuple[int, int], ActivityResult],
        cp_score: dict[tuple[int, int], float],
    ) -> None:
        ws = wb.create_sheet(title=f"Checkpoint {index}")
        headers = [
            "Team",
            "Versus Pair",
            "Match Result",
            "Extra Shots",
            "Vomit penalty (-)",
            "Not drinking penalty (-)",
            "Notes",
            "Total Checkpoint",
        ]
        ws.append(headers)

        for team in teams:
            key = (team.id, checkpoint.id)
            result = cp_result.get(key)
            total = cp_score.get(key, 0.0)
            if result is None:
                # Team never recorded a result at this checkpoint.
                ws.append([team.name, opponent_of.get(team.id, ""), "", 0, 0, 0, "", total])
                continue
            ws.append(
                [
                    team.name,
                    opponent_of.get(team.id, ""),
                    total,  # Match Result mirrors the final checkpoint points.
                    int(result.extra_shots or 0),
                    self._penalty(result, _VOMIT_KEY),
                    self._penalty(result, _NOT_DRINKING_KEY),
                    self._notes(result),
                    total,
                ]
            )

        self._finalize(ws, len(headers), body_center_from_col=3, skip_center_cols={7})

    def _finalize(
        self,
        ws: Worksheet,
        ncols: int,
        *,
        body_center_from_col: int,
        skip_center_cols: set[int] | None = None,
    ) -> None:
        skip = skip_center_cols or set()
        _style_header(ws, ncols)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ncols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = _BODY_FONT
                if col >= body_center_from_col and col not in skip:
                    cell.alignment = _CENTER
        ws.freeze_panes = "A2"
        _autosize(ws, ncols)
