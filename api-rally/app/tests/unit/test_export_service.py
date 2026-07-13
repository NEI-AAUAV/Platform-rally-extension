"""Unit tests for ExportService workbook construction.

These exercise the layout/aggregation logic in isolation: the three DB reads
(`_teams`, `_checkpoints`, `_results`) are stubbed so no database is needed,
and the resulting workbook is parsed back with openpyxl.
"""
from io import BytesIO
from types import SimpleNamespace

import openpyxl
import pytest

from app.services.export_service import ExportService


def _team(tid: int, name: str, group: int | None = None):
    return SimpleNamespace(id=tid, name=name, versus_group_id=group)


def _cp(cid: int, order: int):
    return SimpleNamespace(id=cid, order=order, event_id=1)


def _result(
    team_id: int,
    checkpoint_id: int,
    *,
    final_score: float = 6.0,
    extra_shots: int = 0,
    penalties: dict | None = None,
    notes: str = "",
    completed: bool = True,
):
    return SimpleNamespace(
        team_id=team_id,
        activity=SimpleNamespace(checkpoint=SimpleNamespace(id=checkpoint_id)),
        is_completed=completed,
        final_score=final_score,
        extra_shots=extra_shots,
        penalties=penalties or {},
        result_data={"notes": notes} if notes else {},
        team_vs_result="win",
    )


def _service(teams, checkpoints, results) -> ExportService:
    svc = ExportService.__new__(ExportService)

    async def _t(_event_id):
        return teams

    async def _c(_event_id):
        return checkpoints

    async def _r(_event_id):
        return results

    svc._teams = _t  # type: ignore[method-assign]
    svc._checkpoints = _c  # type: ignore[method-assign]
    svc._results = _r  # type: ignore[method-assign]
    return svc


async def _build(teams, checkpoints, results):
    data = await _service(teams, checkpoints, results).build_workbook(1)
    return openpyxl.load_workbook(BytesIO(data))


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


@pytest.mark.asyncio
async def test_sheet_names_overall_plus_one_per_checkpoint():
    wb = await _build(
        [_team(1, "A")],
        [_cp(10, 1), _cp(11, 2), _cp(12, 3)],
        [],
    )
    assert wb.sheetnames == ["Overall", "Checkpoint 1", "Checkpoint 2", "Checkpoint 3"]


@pytest.mark.asyncio
async def test_overall_headers_and_totals():
    teams = [_team(1, "A", 5), _team(2, "B", 5)]
    cps = [_cp(10, 1), _cp(11, 2)]
    results = [
        _result(1, 10, final_score=6.0),
        _result(1, 11, final_score=3.0),
        _result(2, 10, final_score=4.0),
    ]
    wb = await _build(teams, cps, results)
    rows = _rows(wb["Overall"])

    assert rows[0] == ("Team", "Versus Pair", "Checkpoint 1", "Checkpoint 2", "Total Points")
    # Team A: 6 + 3 = 9, opponent B
    assert rows[1] == ("A", "B", 6, 3, 9)
    # Team B: 4 + 0 = 4, opponent A
    assert rows[2] == ("B", "A", 4, 0, 4)


@pytest.mark.asyncio
async def test_incomplete_result_excluded_from_score():
    teams = [_team(1, "A")]
    cps = [_cp(10, 1)]
    results = [_result(1, 10, final_score=9.0, completed=False)]
    wb = await _build(teams, cps, results)

    # Not completed -> contributes 0 to totals. (openpyxl reads an empty
    # string cell back as None.)
    assert _rows(wb["Overall"])[1] == ("A", None, 0, 0)


@pytest.mark.asyncio
async def test_checkpoint_sheet_columns():
    teams = [_team(1, "A", 5), _team(2, "B", 5)]
    cps = [_cp(10, 1)]
    results = [
        _result(
            1, 10,
            final_score=6.0,
            extra_shots=5,
            penalties={"vomit": 5, "not_drinking": 2},
            notes="Objeto: pasta",
        ),
    ]
    wb = await _build(teams, cps, results)
    rows = _rows(wb["Checkpoint 1"])

    assert rows[0] == (
        "Team", "Versus Pair", "Match Result", "Extra Shots",
        "Vomit penalty (-)", "Not drinking penalty (-)", "Notes", "Total Checkpoint",
    )
    # Team A with a result row.
    assert rows[1] == ("A", "B", 6, 5, 5, 2, "Objeto: pasta", 6)
    # Team B has no result -> zeroed row; blank match result & notes read
    # back as None from openpyxl.
    assert rows[2] == ("B", "A", None, 0, 0, 0, None, 0)


@pytest.mark.asyncio
async def test_unpaired_team_has_empty_versus():
    teams = [_team(1, "Solo", None)]
    wb = await _build(teams, [_cp(10, 1)], [])
    assert _rows(wb["Overall"])[1][1] is None  # Versus Pair blank (empty -> None)


@pytest.mark.asyncio
async def test_malformed_penalty_value_defaults_to_zero():
    teams = [_team(1, "A")]
    cps = [_cp(10, 1)]
    results = [_result(1, 10, penalties={"vomit": "oops"})]
    wb = await _build(teams, cps, results)
    # Non-int penalty coerces to 0 rather than raising.
    assert _rows(wb["Checkpoint 1"])[1][4] == 0
