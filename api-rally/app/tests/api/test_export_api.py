"""API tests for the event results export endpoint, against real Postgres."""

from io import BytesIO

import openpyxl
import pytest

from app.models.activity import Activity, ActivityResult, RallyEvent
from app.models.checkpoint import CheckPoint
from app.models.team import Team

pytestmark = pytest.mark.asyncio

_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def _seed_event(pg_session) -> dict:
    """Create an event with two versus teams, two checkpoints and results."""
    event = RallyEvent(name="Rally Export", event_type="rally_tascas", is_current=True)
    pg_session.add(event)
    await pg_session.commit()
    await pg_session.refresh(event)

    cp1 = CheckPoint(name="CP1", order=1, event_id=event.id)
    cp2 = CheckPoint(name="CP2", order=2, event_id=event.id)
    team_a = Team(name="Alpha", access_code="EXP-A001", event_id=event.id, versus_group_id=1)
    team_b = Team(name="Bravo", access_code="EXP-B001", event_id=event.id, versus_group_id=1)
    pg_session.add_all([cp1, cp2, team_a, team_b])
    await pg_session.commit()
    for obj in (cp1, cp2, team_a, team_b):
        await pg_session.refresh(obj)

    act1 = Activity(
        name="Match CP1",
        activity_type="TeamVsActivity",
        checkpoint_id=cp1.id,
        event_id=event.id,
        config={},
    )
    act2 = Activity(
        name="Match CP2",
        activity_type="TeamVsActivity",
        checkpoint_id=cp2.id,
        event_id=event.id,
        config={},
    )
    pg_session.add_all([act1, act2])
    await pg_session.commit()
    for obj in (act1, act2):
        await pg_session.refresh(obj)

    pg_session.add_all(
        [
            ActivityResult(
                activity_id=act1.id,
                team_id=team_a.id,
                is_completed=True,
                final_score=6.0,
                extra_shots=5,
                penalties={"not_drinking": 1},
                result_data={"result": "win", "notes": "Objeto: pasta"},
                team_vs_result="win",
            ),
            ActivityResult(
                activity_id=act1.id,
                team_id=team_b.id,
                is_completed=True,
                final_score=3.0,
                extra_shots=0,
                penalties={"vomit": 5},
                result_data={"result": "lose"},
                team_vs_result="lose",
            ),
            ActivityResult(
                activity_id=act2.id,
                team_id=team_a.id,
                is_completed=True,
                final_score=3.0,
                result_data={"result": "lose"},
                team_vs_result="lose",
            ),
        ]
    )
    await pg_session.commit()

    return {"event": event, "team_a": team_a, "team_b": team_b}


async def test_export_requires_admin(pg_session, pg_client, as_user):
    seed = await _seed_event(pg_session)
    resp = pg_client.get(f"/api/rally/v1/events/{seed['event'].id}/export")
    assert resp.status_code == 403


def test_export_event_not_found(pg_client, as_admin):
    resp = pg_client.get("/api/rally/v1/events/999999/export")
    assert resp.status_code == 404


async def test_export_returns_xlsx_with_expected_sheets(pg_session, pg_client, as_admin):
    seed = await _seed_event(pg_session)

    resp = pg_client.get(f"/api/rally/v1/events/{seed['event'].id}/export")

    assert resp.status_code == 200
    assert resp.headers["content-type"] == _XLSX_MEDIA_TYPE
    assert "attachment" in resp.headers["content-disposition"]
    assert "Rally_Export_results.xlsx" in resp.headers["content-disposition"]

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    assert wb.sheetnames == ["Overall", "Checkpoint 1", "Checkpoint 2"]


async def test_export_overall_scores_and_versus(pg_session, pg_client, as_admin):
    seed = await _seed_event(pg_session)

    resp = pg_client.get(f"/api/rally/v1/events/{seed['event'].id}/export")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    rows = list(wb["Overall"].iter_rows(values_only=True))

    assert rows[0] == ("Team", "Versus Pair", "Checkpoint 1", "Checkpoint 2", "Total Points")
    # Teams ordered by name: Alpha then Bravo.
    assert rows[1] == ("Alpha", "Bravo", 6, 3, 9)
    assert rows[2] == ("Bravo", "Alpha", 3, 0, 3)


async def test_export_checkpoint_detail_columns(pg_session, pg_client, as_admin):
    seed = await _seed_event(pg_session)

    resp = pg_client.get(f"/api/rally/v1/events/{seed['event'].id}/export")
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    rows = list(wb["Checkpoint 1"].iter_rows(values_only=True))

    assert rows[0] == (
        "Team",
        "Versus Pair",
        "Match Result",
        "Extra Shots",
        "Vomit penalty (-)",
        "Not drinking penalty (-)",
        "Notes",
        "Total Checkpoint",
    )
    assert rows[1] == ("Alpha", "Bravo", 6, 5, 0, 1, "Objeto: pasta", 6)
    assert rows[2] == ("Bravo", "Alpha", 3, 0, 5, 0, None, 3)
